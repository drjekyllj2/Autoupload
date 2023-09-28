import pandas as pd 
import openpyxl
import csv
import os
from collections import Counter
import configparser as config
import logging
import datetime
from sendmail import send_mail

 

config_out= config.ConfigParser()
config_out.read ("Config.ini") 
current_directory = os.getcwd()
current_dir = current_directory.replace("\\","/")
excel_path =config_out.get('file_name','combined_file')
consolidated_log =current_dir + config_out.get('log_name','consolidated_log')
file_tosend = current_dir + '/updated_excel_file.csv'
email_to =  config_out.get('email_config','netsuite_mailto')
ccmail_to = config_out.get('email_config','log_ccmailto')
log_filename = config_out.get('log_name','download_log')
log_toemail = current_dir + log_filename
#excel_path = "../Combine.xlsx"
merged_file = config_out.get('file_name','deduped_file')#"merged_output.xlsx"
logging.basicConfig(
        filename=log_toemail,
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
        ) 
def process_excel():
    try:
        df = pd.read_excel(excel_path)
        # Group by the first column and aggregate other columns by concatenating them
        merged_df = df.groupby('Entry Name').agg(lambda x: ', '.join(x.astype(str))).reset_index()
        
        #grouped = df.groupby('Entry Name').agg({'Photo1': ', '.join}).reset_index()
        # Merge the aggregated data back to the original DataFrame
        #merged_df = pd.merge(df, grouped, on='Entry Name', how='left')
        
    
        merged_df.to_excel(merged_file, index=False)
        
        return True
    except Exception as e:
            print(f"Error: {str(e)}") 
            logging.info(f"Error : {str(e)}")
            send_mail(email_to,"Autoupload error/info log","Attached is the autoupload log file.",log_toemail,ccmail_to)     
            logging.shutdown()
            return False
      
def process_file():
    try:
        wb = openpyxl.load_workbook(merged_file)
        sheet = wb["Sheet1"]
        data_rows = []
        column_indices=[]
        #column_indices = [14, 15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37]
        for indices in range(14,38):
            column_indices.append(indices)

    # Iterate through rows to insert new cells in each row at the specified column indices
        for row in sheet.iter_rows(min_row=1, values_only=True):
            for column_index in column_indices:
                sheet.cell(row=1, column=column_index, value=None)  # Insert an empty cell
                
        # Set headers for the new columns
        photo_no = 9
        for icol in range(0,24):
            sheet.cell(row=1, column=column_indices[icol], value=(f'Photo{photo_no}'))
            photo_no = photo_no +1
            
        # sheet.cell(row=1, column=column_indices[0], value='Photo9')
        # sheet.cell(row=1, column=column_indices[1], value='Photo10') 
        # sheet.cell(row=1, column=column_indices[2], value='Photo11')
        # sheet.cell(row=1, column=column_indices[3], value='Photo12')
        # sheet.cell(row=1, column=column_indices[4], value='Photo13')
        # sheet.cell(row=1, column=column_indices[5], value='Photo14')
        # sheet.cell(row=1, column=column_indices[6], value='Photo15')
        # sheet.cell(row=1, column=column_indices[7], value='Photo16')
        # sheet.cell(row=1, column=column_indices[8], value='Photo17')
        # sheet.cell(row=1, column=column_indices[9], value='Photo18')
        # sheet.cell(row=1, column=column_indices[10], value='Photo19')
        # sheet.cell(row=1, column=column_indices[11], value='Photo20')
        # sheet.cell(row=1, column=column_indices[12], value='Photo21')
        # sheet.cell(row=1, column=column_indices[13], value='Photo22')
        # sheet.cell(row=1, column=column_indices[14], value='Photo23')
        # sheet.cell(row=1, column=column_indices[15], value='Photo24')
        # sheet.cell(row=1, column=column_indices[16], value='Photo25')
        # sheet.cell(row=1, column=column_indices[17], value='Photo26')
        # sheet.cell(row=1, column=column_indices[18], value='Photo27')
        # sheet.cell(row=1, column=column_indices[19], value='Photo28')
        # sheet.cell(row=1, column=column_indices[20], value='Photo29')
        # sheet.cell(row=1, column=column_indices[21], value='Photo30')
        # sheet.cell(row=1, column=column_indices[22], value='Photo31')
        # sheet.cell(row=1, column=column_indices[23], value='Photo32')
        i=2
        col=6
        data_col=[]
        data_tocsv=[]
        for   row in sheet.iter_rows(min_row=2, values_only=True):
            #print (f"print rows {row[1]}")
            input_string=row[1]
            if row[1].find("nan")!=-1 :
                continue
            print (f"this is row 1 value {row[1]}")
            result_array = input_string.split(',')
            for colindx in range( 5, 13):
                col_string =row[colindx]
                colvalue_list = col_string.split(',') 
                # colvalresult = colvalresult + colvalue_list
                
                for col_list in colvalue_list:
                    data_col.append(col_list)
                    print (f" {row[0]}  column {colindx}  photo  {col_list}  ")
                    #print (data_col)
            int_list = [float(x) for x in result_array]
            data_length=len(data_col)
            
            print(f"{row[0]} length {data_length}")
            #print ( f" old {input_string}  new result  {result_array} ")
            qty = sum(int_list)
            if row[2].find("nan") !=-1:
                sheet.cell(row=i, column=3, value="")
            if row[3].find("nan") !=-1:
                sheet.cell(row=i, column=4, value="")    
            sheet.cell(row=i, column=2, value=qty)
            
            split_tags = row[4].split(',')
            approved_cntr=0
            output_string=""
            strip_space=[word.strip() for word in split_tags]
            distinct_tags= list(Counter(strip_space).keys())
            print(f"distinct tags {distinct_tags}")
            tag_count = len(distinct_tags)
            cntr=1
            for itag in distinct_tags:
                if cntr < tag_count:
                    itag=itag + " ,"
                output_string=output_string  + itag
                cntr = cntr+1
            sheet.cell(row=i,column=5,value = output_string)
            for data in data_col:
                if data.find("nan")!=-1:
                    data=""
                if col>=38:
                    break
                else:
                    sheet.cell(row=i, column=col, value=data)
                col=col+1
            data_col=[]
            col=6    
            # for cell in row:
            #     #data_rows.append(row)
            #     input_string = cell
                #result_array = input_string.split(',')
            i=i+1
            
        wb.save('updated_excel_file.xlsx')
        wb.close() 
        del wb
        
        updated_excel = 'updated_excel_file.xlsx' 
        wbcsv = openpyxl.load_workbook(updated_excel)
        sheet_csv = wbcsv.active
        
        for rows in sheet_csv.iter_rows(values_only=True):
            data_tocsv.append(rows)
           
        with open("updated_excel_file.csv", "w", newline="") as csvfile:
            csvwriter = csv.writer(csvfile)
            csvwriter.writerows(data_tocsv)
        wbcsv.close()        
        # Close the Excel file
        
        send_mail(email_to,"CSV Report uploaded to Netsuite", "The CSV report has been sent to netsuite. File name : updated_excel_file.csv" ,file_tosend,ccmail_to)    
        logging.shutdown()
    except Exception as e:
            print(f"Error: {str(e)}") 
            logging.info(f"Error : {str(e)}")
            send_mail(email_to,"Autoupload error/info log","Attached is the autoupload log file.",log_toemail,ccmail_to)     
            logging.shutdown()    
    # Print the data rows
        # for row in data_rows:
        #     print(row)
    #process_excel()
    #process_file()    