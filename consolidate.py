import pandas as pd 
import openpyxl
import os
import csv
import configparser as config
import logging as log
import datetime
from sendmail import send_mail
 

#url = "https://s3.us-west-2.amazonaws.com/data.prod.sortly.com/400902cb56115ead1b0eac5a/csvs/963ec423dbe5be4f6ce07340cbeb8bae0570b675.original.csv?response-content-disposition=attachment&X-Amz-Algorithm=AWS4-HMAC-SHA256&X-Amz-Credential=AKIAINAWTSTHEUFHL3VA%2F20230907%2Fus-west-2%2Fs3%2Faws4_request&X-Amz-Date=20230907T062606Z&X-Amz-Expires=86400&X-Amz-SignedHeaders=host&X-Amz-Signature=56a8abdf6f4b1fa2f8bc622a100f595df453cfdd6407b2ed75e6a144eb9d8b7b"  # Replace with the URL of the file you want to download


current_directory = os.getcwd()
current_dir = current_directory.replace("\\","/")
config_out= config.ConfigParser()
config_out.read ("Config.ini")
log_filename=config_out.get('log_name','download_log') 
consolidated_log = current_dir + log_filename
email_to =  config_out.get('email_config','log_mailto')
ccmail_to = config_out.get('email_config','log_ccmailto')
mergedrecords =config_out.get('file_name','combined_file')
log.basicConfig(
        filename=consolidated_log,
        level=log.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
        )   
def consolidate_data():
    try:
        
        filename =str( datetime.date.today() )  
        save_path = config_out.get('file_directory','fromemail')
        # Replace with the desired local file path and name
        
        email_csv =current_dir + "/" + filename + "-" +config_out.get('file_name','mail_csvfilename')  # "../downloaded-file.csv"
        from_email =current_dir + "/" + filename + "-downloaded-file.xlsx"
      
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        

        # Open and read the CSV file
        with open(email_csv, 'r', newline='') as csv_file:
            csv_reader = csv.reader(csv_file)
            for row in csv_reader:
                worksheet.append(row)

        # Save the Excel workbook as an XLSX file
        workbook.save(from_email)
        workbook.close()

        
        excel_path =mergedrecords# "../Combine.xlsx"
        
        wb_email = openpyxl.load_workbook(from_email)
        wb_consolidate = openpyxl.load_workbook(excel_path)
        sheet_email = wb_email.active
        sheet_consolidate=wb_consolidate["Temp"]
        data_rows = []
        append_firstrowcol = sheet_consolidate.max_row + 1
        
        email_col=0 
        
        for irow in sheet_email.iter_rows(min_row=2, values_only=True):
            
            entry_name = irow[0]
            
            qty = int(irow[10]) if irow[10] != None else ""
            min_level=irow[11]
            notes=irow[14]
            tag=irow[15]
            # photo1=row[21]
            # photo2=row[22]
            # photo3=row[23]
            # photo4=row[24]
            # photo5=row[25]
            # photo6=row[26]
            # photo7=row[27]
            # photo8=row[28]
            sheet_consolidate.cell(row= append_firstrowcol, column=1, value=entry_name)
            sheet_consolidate.cell(row=append_firstrowcol, column=2, value=qty)
            sheet_consolidate.cell(row=append_firstrowcol, column=3, value=min_level)
            sheet_consolidate.cell(row=append_firstrowcol, column=4, value=notes)
            sheet_consolidate.cell(row=append_firstrowcol, column=5, value=tag)
            
            email_col=21
            for col_indx in range(6,14):
                sheet_consolidate.cell(row=append_firstrowcol, column=col_indx, value=irow[email_col])
                email_col = email_col +1
            append_firstrowcol=append_firstrowcol+1
            
        wb_consolidate.save(excel_path)
        wb_consolidate.close()
        wb_email.close() 
        log.info("Records from downloaded file successfully merged to file " + excel_path)
        send_mail(email_to,"Consolidated log","This is a consolidated log",consolidated_log,ccmail_to)
        log.shutdown()
    except FileNotFoundError as e:
           log.info(f"Error : {str(e)}")
           print (f"Error: {str(e)} ")
           log.shutdown()
           send_mail(email_to,"Consolidated Error log","There's an error on consolidating/mering file.",consolidated_log,ccmail_to)    
    except Exception as e:
            print(f"Error: {str(e)}") 
            log.info(f"Error : {str(e)}")
            log.shutdown()
            send_mail(email_to,"Consolidated Error log","There's an error on consolidating/mering file.",consolidated_log,ccmail_to)
            
            
            